'''
Created on 23-Jun-2026

@author: Vivek
'''
from openpyxl import load_workbook

# 1. Loading the workbook
filename = r"D:\TestData\OrangeHRMLoginPageDDT.xlsx"
my_workbook = load_workbook(filename)

# 2. Set the active sheet
sheet1 = my_workbook["Sheet1"]

# 3. Get total no. of cols and rows
total_cols = sheet1.max_column
print('total_cols:', total_cols)
total_rows = sheet1.max_row
print('total_rows:', total_rows)
print("==========================")

# 4. Write into result column
for i in range(2, total_rows+1):
    
    sheet1.cell(i,5).value = "passed"
    
my_workbook.save(filename)
    